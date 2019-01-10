import openpyxl

# TODO:
# 1. Specify if there is a header row and skip it
# 2. If header row, keep the column names in order to be used as headers
# 3. Check if column names can be used as associative array subscripts

class XlsxReader:

    def __init__(self, filename):
        workbook = openpyxl.load_workbook(filename + '.xlsx')
        self.sheet = workbook.active

    def display(self):
        max_row = self.sheet.max_row

        for i in range(1, max_row+1):
            columns = self.sheet.max_column
            for j in range(1, columns+1):
                value = self.sheet.cell(row=i,column=j).value
                print(value, end=", ")
            print()


def main():
    excelReader = XlsxReader('100-contacts')
    excelReader.display()


if __name__== "__main__":
    main()
